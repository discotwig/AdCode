import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def _tracked_files():
    result = subprocess.run(
        ["git", "ls-files"],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return [ROOT / line for line in result.stdout.splitlines() if line]


def _tracked_text_files():
    for path in _tracked_files():
        if not path.exists():
            continue
        if path.suffix.lower() in {".xlsx", ".png", ".jpg", ".jpeg", ".gif", ".ico"}:
            continue
        try:
            yield path, path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue


def test_no_personal_deployment_details_in_tracked_text_files():
    personal_domain = "ryanbishop" + "." + "me"
    banned = [
        personal_domain,
        "bishopryant" + "@gmail.com",
        "traffic@" + personal_domain,
        "https://api." + personal_domain + "/inbound",
    ]

    offenders = []
    for path, text in _tracked_text_files():
        for value in banned:
            if value in text:
                offenders.append(f"{path.relative_to(ROOT)} contains {value}")

    assert offenders == []


def test_no_known_real_demo_account_or_page_ids_in_tracked_files():
    real_account = re.compile(r"act_" + "366643171" + "197739")
    real_page = re.compile(r"(?<!\d)" + "102066921" + "653934" + r"(?!\d)")

    offenders = []
    for path, text in _tracked_text_files():
        if real_account.search(text) or real_page.search(text):
            offenders.append(str(path.relative_to(ROOT)))

    for path in [
        ROOT / "docs/demo/client_brief_may2026.xlsx",
        ROOT / "docs/demo/client_campaign_tracker.xlsx",
        ROOT / "docs/demo/client_campaign_tracker_v2.xlsx",
    ]:
        wb = load_workbook(path, data_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = "" if cell.value is None else str(cell.value)
                    if real_account.search(value) or real_page.search(value):
                        offenders.append(f"{path.relative_to(ROOT)}:{ws.title}!{cell.coordinate}")

    assert offenders == []


def test_private_workspace_and_local_config_are_not_tracked():
    tracked = {path.relative_to(ROOT).as_posix() for path in _tracked_files()}

    assert not any(path.startswith("customers/") for path in tracked)
    assert ".claude/settings.local.json" not in tracked


def test_no_private_env_or_credential_files_are_tracked():
    forbidden_names = {".env", ".env.local", ".env.production"}
    forbidden_fragments = ("credential", "credentials", "secret", "token", "private")

    offenders = []
    for path in _tracked_files():
        rel = path.relative_to(ROOT).as_posix()
        lower = rel.lower()
        if path.name in forbidden_names:
            offenders.append(rel)
        elif any(fragment in lower for fragment in forbidden_fragments) and rel != ".env.example":
            offenders.append(rel)

    assert offenders == []


def test_current_security_docs_use_strict_iac_tool_names():
    security = (ROOT / "SECURITY.md").read_text(encoding="utf-8")

    assert "plan_stack" in security
    assert "apply_stack" in security
    assert "import_resource" in security
    assert "plan_campaigns" not in security
    assert "apply_campaigns" not in security
    assert "import_adsets" not in security
