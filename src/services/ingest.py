import argparse
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

SCHEMA_PATH = Path(__file__).parent.parent.parent / "schemas" / "campaign.schema.json"


@dataclass
class Ambiguity:
    field: str
    sheet: str
    cell_ref: str
    raw_value: str
    question: str


@dataclass
class IngestionResult:
    campaigns: list[dict] = field(default_factory=list)
    ambiguities: list[Ambiguity] = field(default_factory=list)
    confidence: float = 1.0


def read_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                data_rows.append(dict(zip(headers, row)))
        result[sheet_name] = {"headers": headers, "rows": data_rows, "raw_rows": rows}
    return result


_INGEST_PROMPT = """\
You are an ad trafficking assistant. Your job is to extract Facebook ad campaign definitions
from an Excel file and convert them into structured JSON matching the schema below.

The Excel file is non-standard and may have merged cells, colour-coded rows, multi-sheet layouts,
or non-standard column names. Use your best judgement to interpret the content.

Campaign JSON schema (required fields only, refer to this for structure):
{schema}

Excel file contents:
{excel_data}

Instructions:
1. Extract all distinct ad campaigns you can identify.
2. For each campaign, create a complete JSON object matching the schema.
3. Use "PAUSED" as the default status for all objects unless the Excel clearly indicates otherwise.
4. If a required field is ambiguous or missing, use your best guess and record it as an ambiguity.
5. The account_id should be taken from the Excel if present; otherwise use "act_000000000".

Respond with a JSON object with two keys:
- "campaigns": array of campaign JSON objects matching the schema structure (just the array of campaign objects, not the top-level wrapper)
- "ambiguities": array of ambiguity objects, each with:
  - "field": the JSON path of the ambiguous field
  - "sheet": the Excel sheet name where the issue was found
  - "cell_ref": the cell reference (e.g. "B5") if identifiable, otherwise ""
  - "raw_value": the raw value from Excel that was ambiguous
  - "question": what you are unsure about and what human review should verify
- "confidence": a float from 0.0 to 1.0 representing your overall confidence in the extraction

Return only the JSON object, no other text."""


def extract_campaigns(excel_data: dict, ai_client) -> IngestionResult:
    with open(SCHEMA_PATH) as f:
        schema = json.load(f)

    prompt = _INGEST_PROMPT.format(
        schema=json.dumps(schema, indent=2),
        excel_data=json.dumps(excel_data, indent=2, default=str),
    )

    response = ai_client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = response.content[0].text.strip()

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        logger.error("Ingestion AI returned non-JSON response: %s", e)
        return IngestionResult(confidence=0.0)

    campaigns = parsed.get("campaigns", [])
    confidence = float(parsed.get("confidence", 0.5))
    ambiguities = []
    for item in parsed.get("ambiguities", []):
        try:
            ambiguities.append(Ambiguity(
                field=item.get("field", ""),
                sheet=item.get("sheet", ""),
                cell_ref=item.get("cell_ref", ""),
                raw_value=str(item.get("raw_value", "")),
                question=item.get("question", ""),
            ))
        except (KeyError, TypeError, AttributeError) as e:
            logger.warning("Skipping malformed ambiguity: %s", e)

    return IngestionResult(campaigns=campaigns, ambiguities=ambiguities, confidence=confidence)


def format_ambiguity_report(result: IngestionResult) -> str:
    if not result.ambiguities:
        return f"No ambiguities found. Confidence: {result.confidence:.0%}. {len(result.campaigns)} campaign(s) extracted."

    lines = [
        f"Ingestion complete. {len(result.campaigns)} campaign(s) extracted. "
        f"Confidence: {result.confidence:.0%}.",
        f"{len(result.ambiguities)} ambiguity/ambiguities require human review:",
        "",
    ]
    for i, amb in enumerate(result.ambiguities, 1):
        lines.append(f"{i}. [{amb.sheet}] {amb.field}")
        if amb.cell_ref:
            lines.append(f"   Cell: {amb.cell_ref}  Raw value: {amb.raw_value!r}")
        else:
            lines.append(f"   Raw value: {amb.raw_value!r}")
        lines.append(f"   Question: {amb.question}")
        lines.append("")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Ingest an Excel file and extract campaign JSON.")
    parser.add_argument("excel_file", help="Path to the Excel file")
    parser.add_argument("--output", required=True, help="Output JSON file path")
    args = parser.parse_args()

    import os
    import anthropic
    from dotenv import load_dotenv
    load_dotenv()

    ai_client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    excel_data = read_excel(args.excel_file)
    result = extract_campaigns(excel_data, ai_client)

    print(format_ambiguity_report(result))

    if result.campaigns:
        account_id = "act_000000000"
        output = {"account_id": account_id, "campaigns": result.campaigns}
        with open(args.output, "w") as f:
            json.dump(output, f, indent=2)
        print(f"\nCampaign JSON written to {args.output}")
        print("Review the ambiguities above before committing.")
    else:
        print("No campaigns extracted. Check the Excel file and try again.")


if __name__ == "__main__":
    main()
